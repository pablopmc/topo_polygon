from __future__ import annotations

from pathlib import Path
import openpyxl

from core.reports.excel_report import generate_excel_report


def test_excel_report_generation(tmp_path: Path) -> None:
    output_file = tmp_path / "test_report.xlsx"
    
    project = {
        "name": "Projeto Teste Excel",
        "client": "Cliente de Teste S/A",
        "location": "Avenida Brasil, 123",
        "survey_description": "Descrição de teste para exportação excel",
        "surveyor": "Engenheiro Testador",
        "institution": "Universidade de Testes",
        "survey_date": "2026-06-21",
        "coordinate_system": "UTM WGS84 22S",
        "reference_point": "Marco Zero",
        "azimute_inicial": 45.5,
    }
    
    leituras = [
        {"seq": 1, "point": "V1", "graus": 90, "minutos": 0, "segundos": 0.0, "distancia": 100.0, "observacao": "Obs 1"},
        {"seq": 2, "point": "V2", "graus": 90, "minutos": 0, "segundos": 0.0, "distancia": 100.0, "observacao": ""},
    ]
    
    calculos = [
        {"seq": 1, "point": "V1", "azimute": 45.5, "rumo": "N 45 30 00 E", "quadrante": "NE", "dx": 71.325, "dy": 70.091, "distancia": 100.0},
    ]
    
    coordenadas = [
        {"seq": 1, "point": "V1", "x": 0.0, "y": 0.0, "x_corrigido": 0.0, "y_corrigido": 0.0},
    ]
    
    generate_excel_report(
        filename=output_file,
        project=project,
        leituras=leituras,
        calculos=calculos,
        coordenadas=coordenadas,
        area=124665.7958,
        perimeter=852.123,
        precision=12500.0,
    )
    
    assert output_file.exists()
    
    # Load and inspect the generated workbook using openpyxl
    wb = openpyxl.load_workbook(output_file)
    
    # Check sheets exist
    assert "Leituras" in wb.sheetnames
    assert "Cálculos" in wb.sheetnames
    assert "Coordenadas" in wb.sheetnames
    assert "Área" in wb.sheetnames
    assert "Resumo" in wb.sheetnames
    
    # Inspect Leituras values
    ws_leituras = wb["Leituras"]
    assert ws_leituras.cell(1, 1).value == "Seq"
    assert ws_leituras.cell(2, 1).value == 1
    assert ws_leituras.cell(2, 2).value == "V1"
    assert ws_leituras.cell(2, 6).value == 100.0
    assert ws_leituras.cell(2, 7).value == "Obs 1"
    
    # Inspect Cálculos values
    ws_calculos = wb["Cálculos"]
    assert ws_calculos.cell(1, 3).value == "Azimute"
    assert ws_calculos.cell(2, 3).value == 45.5
    assert ws_calculos.cell(2, 6).value == 71.325
    
    # Inspect Coordenadas values
    ws_coordenadas = wb["Coordenadas"]
    assert ws_coordenadas.cell(1, 5).value == "X Corrigido"
    assert ws_coordenadas.cell(2, 3).value == 0.0
    
    # Inspect Resumo values
    ws_resumo = wb["Resumo"]
    # Check project metadata label/value
    meta_dict = {}
    for r in range(2, ws_resumo.max_row + 1):
        lbl = ws_resumo.cell(r, 1).value
        val = ws_resumo.cell(r, 2).value
        if lbl:
            meta_dict[lbl] = val
            
    assert meta_dict["Nome do Projeto"] == "Projeto Teste Excel"
    assert meta_dict["Contratante/Cliente"] == "Cliente de Teste S/A"
    assert meta_dict["Localidade/Endereço"] == "Avenida Brasil, 123"
    assert meta_dict["Instituição"] == "Universidade de Testes"
    assert meta_dict["Responsável"] == "Engenheiro Testador"
    assert meta_dict["Sistema de Coordenadas"] == "UTM WGS84 22S"
    assert meta_dict["Ponto de Referência"] == "Marco Zero"
    
    # Check summary values at the bottom
    # We should search for where summary block starts
    found_summary = False
    for row in range(1, ws_resumo.max_row + 1):
        if ws_resumo.cell(row, 1).value == "Área (m²)":
            assert ws_resumo.cell(row, 2).value == "124665.7958"
            found_summary = True
        elif ws_resumo.cell(row, 1).value == "Precisão":
            assert ws_resumo.cell(row, 2).value == "1:12.500"
            
    assert found_summary
