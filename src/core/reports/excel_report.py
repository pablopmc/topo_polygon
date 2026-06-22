from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

Row = Mapping[str, object]

_HEADER_TO_KEY = {
    # Leituras
    "Seq": "seq",
    "Ponto": "point",
    "Grau": "graus",
    "Min": "minutos",
    "Seg": "segundos",
    "Distância": "distancia",
    "Observação": "observacao",
    
    # Cálculos
    "Azimute": "azimute",
    "Rumo": "rumo",
    "Quadrante": "quadrante",
    "ΔX": "dx",
    "ΔY": "dy",
    
    # Coordenadas
    "X": "x",
    "Y": "y",
    "X Corrigido": "x_corrigido",
    "Y Corrigido": "y_corrigido",
}


def _get_row_value(row: Row, header: str) -> object:
    key = _HEADER_TO_KEY.get(header, header.lower())
    val = row.get(key)
    if val is None:
        val = row.get(header)
    if val is None:
        val = row.get(header.lower())
    return val


def _format_value(value: object) -> object:
    return "" if value is None else value


def _style_header(row_cells) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in row_cells:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _adjust_column_widths(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        width = min(50, max(10, max_length + 2))
        sheet.column_dimensions[column].width = width


def _write_rows(sheet, headers: Sequence[str], rows: Iterable[Row]) -> None:
    sheet.append(list(headers))
    _style_header(sheet[1])
    for row in rows:
        sheet.append([_format_value(_get_row_value(row, column)) for column in headers])
    sheet.freeze_panes = "A2"
    _adjust_column_widths(sheet)


def _write_project_metadata(sheet, project: Row) -> None:
    sheet.append(["Campo", "Valor"])
    _style_header(sheet[1])
    
    name = project.get("name") or project.get("project_name") or "PROJETO"
    client = project.get("client") or ""
    location = project.get("location") or ""
    desc = project.get("survey_description") or project.get("description") or ""
    surveyor = project.get("surveyor") or project.get("author") or ""
    date = project.get("survey_date") or project.get("created_at") or ""
    inst = project.get("institution") or ""
    coordinate_system = project.get("coordinate_system") or ""
    reference_point = project.get("reference_point") or ""
    az_init = project.get("azimute_inicial")
    
    metadata = [
        ("Nome do Projeto", name),
        ("Contratante/Cliente", client),
        ("Localidade/Endereço", location),
        ("Descrição", desc),
        ("Responsável", surveyor),
        ("Instituição", inst),
        ("Data do Levantamento", date),
        ("Sistema de Coordenadas", coordinate_system),
        ("Ponto de Referência", reference_point),
    ]
    
    if az_init is not None:
        from core.calculations.engine import decimal_para_dms
        try:
            g, m, s = decimal_para_dms(float(az_init))
            sinal = "-" if g < 0 or m < 0 or s < 0 else ""
            az_str = f"{sinal}{abs(g)}° {abs(m):02d}' {abs(s):05.2f}\""
        except Exception:
            az_str = str(az_init)
        metadata.append(("Azimute Inicial", az_str))

    for label, value in metadata:
        sheet.append([label, _format_value(value)])
    _adjust_column_widths(sheet)


def _write_summary(sheet, area: float, perimeter: float, precision: float | None) -> None:
    sheet.append(["Resumo", "Valor"])
    _style_header(sheet[1])
    sheet.append(["Área (m²)", f"{area:.4f}"])
    sheet.append(["Perímetro (m)", f"{perimeter:.4f}"])
    
    # Prefix relative precision with 1: if valid
    if precision is None:
        prec_str = ""
    elif precision == float("inf"):
        prec_str = "1:∞"
    else:
        prec_str = f"1:{int(round(precision)):,}".replace(",", ".")
    sheet.append(["Precisão", prec_str])
    _adjust_column_widths(sheet)


def generate_excel_report(
    filename: str | Path,
    project: Row,
    leituras: Sequence[Row],
    calculos: Sequence[Row],
    coordenadas: Sequence[Row],
    area: float,
    perimeter: float,
    precision: float | None = None,
) -> None:
    """
    Salva um workbook Excel contendo planilhas de leituras, cálculos,
    coordenadas, área e resumo.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet_leituras = workbook.create_sheet("Leituras")
    _write_rows(
        sheet_leituras,
        ["Seq", "Ponto", "Grau", "Min", "Seg", "Distância", "Observação"],
        leituras,
    )

    sheet_calculos = workbook.create_sheet("Cálculos")
    _write_rows(
        sheet_calculos,
        ["Seq", "Ponto", "Azimute", "Rumo", "Quadrante", "ΔX", "ΔY", "Distância"],
        calculos,
    )

    sheet_coordenadas = workbook.create_sheet("Coordenadas")
    _write_rows(
        sheet_coordenadas,
        ["Seq", "Ponto", "X", "Y", "X Corrigido", "Y Corrigido"],
        coordenadas,
    )

    sheet_area = workbook.create_sheet("Área")
    sheet_area.append(["Descrição", "Valor"])
    _style_header(sheet_area[1])
    sheet_area.append(["Área (m²)", f"{area:.4f}"])
    sheet_area.append(["Perímetro (m)", f"{perimeter:.4f}"])
    _adjust_column_widths(sheet_area)

    sheet_resumo = workbook.create_sheet("Resumo")
    _write_project_metadata(sheet_resumo, project)
    sheet_resumo.append([])
    _write_summary(sheet_resumo, area, perimeter, precision)

    output_path = Path(filename)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))


from core.reports.exporter_registry import BaseExporter, ExporterRegistry
from core.reports.pdf_report import _row_value
from typing import Any, Mapping


class ExcelExporter(BaseExporter):
    """Adaptador de exportação para formato Excel"""

    def export(
        self,
        filename: str | Path,
        project: Mapping[str, Any],
        calculation_rows: Sequence[Mapping[str, Any]],
        coordinates_rows: Sequence[Mapping[str, Any]],
        sketch_points: Sequence[tuple[float, float]],
        area: float,
        perimeter: float,
        precision: float | None,
        **kwargs: Any,
    ) -> None:
        vertices = kwargs.get("vertices", [])
        resultado = kwargs.get("resultado")
        
        if not vertices or not resultado:
            # Fallback se não fornecido via kwargs (reconstrói a partir dos rows)
            leituras = []
            calculos = []
            coordenadas = []
            for i, row in enumerate(calculation_rows):
                seq = i + 1
                point = _row_value(row, ["Ponto", "point_name", "point"])
                
                # Para leituras, tenta obter do banco se possível, ou usa nulo
                leituras.append({
                    "seq": seq,
                    "point": point,
                    "graus": 0,
                    "minutos": 0,
                    "segundos": 0.0,
                    "distancia": float(_row_value(row, ["Distância (m)", "distancia"]) or 0.0),
                    "observacao": _row_value(row, ["Observação", "observacao", "notes"]),
                })
                
                calculos.append({
                    "seq": seq,
                    "point": point,
                    "azimute": float(_row_value(row, ["Azimute"]) or 0.0),
                    "rumo": _row_value(row, ["Rumo"]),
                    "quadrante": _row_value(row, ["Quadrante"]),
                    "dx": float(_row_value(row, ["Projeção Compensada X (Leste/Oeste)", "dx"]) or 0.0),
                    "dy": float(_row_value(row, ["Projeção Compensada Y (Norte/Sul)", "dy"]) or 0.0),
                    "distancia": float(_row_value(row, ["Distância (m)", "distancia"]) or 0.0),
                })
                
                coordenadas.append({
                    "seq": seq,
                    "point": point,
                    "x": float(_row_value(row, ["Coordenada X"]) or 0.0),
                    "y": float(_row_value(row, ["Coordenada Y"]) or 0.0),
                    "x_corrigido": float(_row_value(row, ["Coordenada X"]) or 0.0),
                    "y_corrigido": float(_row_value(row, ["Coordenada Y"]) or 0.0),
                })
        else:
            leituras = [
                {
                    "seq": v.sequence,
                    "point": v.point_name,
                    "graus": v.graus,
                    "minutos": v.minutos,
                    "segundos": v.segundos,
                    "distancia": v.distancia,
                    "observacao": v.observacao,
                }
                for v in vertices
            ]
            calculos = [
                {
                    "seq": i + 1,
                    "point": v.point_name,
                    "azimute": resultado.azimutes[i],
                    "rumo": resultado.rumos[i],
                    "quadrante": resultado.quadrantes[i],
                    "dx": resultado.deltas_x[i],
                    "dy": resultado.deltas_y[i],
                    "distancia": v.distancia,
                }
                for i, v in enumerate(vertices)
            ]
            coordenadas = [
                {
                    "seq": i + 1,
                    "point": v.point_name,
                    "x": resultado.coordenadas[i][0],
                    "y": resultado.coordenadas[i][1],
                    "x_corrigido": resultado.coordenadas_corrigidas[i][0],
                    "y_corrigido": resultado.coordenadas_corrigidas[i][1],
                }
                for i, v in enumerate(vertices)
            ]

        generate_excel_report(
            filename,
            project=project,
            leituras=leituras,
            calculos=calculos,
            coordenadas=coordenadas,
            area=area,
            perimeter=perimeter,
            precision=precision,
        )


# Registrar exportador no registry global
ExporterRegistry.register(
    format_id="EXCEL",
    display_name="Excel",
    file_filter="Arquivos Excel (*.xlsx)",
    default_ext="xlsx",
    default_name="relatorio.xlsx",
    icon_name="SP_FileDialogDetailedView",
    exporter_class=ExcelExporter,
)